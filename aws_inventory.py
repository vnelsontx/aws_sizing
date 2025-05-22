
import boto3
from botocore.exceptions import ClientError
from collections import defaultdict
import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import datetime

ROLE_NAME = "OrganizationAccountAccessRole"
REGIONS = [
    'us-east-1', 'us-east-2', 'us-west-1', 'us-west-2',
    'us-gov-west-1', 'us-gov-east-1'
]

org_client = boto3.client('organizations')
sts_client = boto3.client('sts')
accounts = org_client.list_accounts()['Accounts']

results = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
summary = []

def assume_role(account_id):
    try:
        response = sts_client.assume_role(
            RoleArn=f"arn:aws:iam::{account_id}:role/{ROLE_NAME}",
            RoleSessionName="InventorySession"
        )
        creds = response['Credentials']
        return dict(
            aws_access_key_id=creds['AccessKeyId'],
            aws_secret_access_key=creds['SecretAccessKey'],
            aws_session_token=creds['SessionToken']
        )
    except ClientError as e:
        print(f"[!] Failed to assume role in {account_id}: {e}")
        return None

def collect_inventory(account_id, creds):
    for region in REGIONS:
        try:
            ec2 = boto3.client('ec2', region_name=region, **creds)
            volumes = ec2.describe_volumes()['Volumes']
            total_size = sum(v['Size'] for v in volumes)
            results[account_id][region]['EC2'] = [{'VolumeId': v['VolumeId'], 'Size': v['Size'], 'Tags': v.get('Tags', [])} for v in volumes]
            summary.append([account_id, region, 'EC2', len(volumes), total_size])
        except Exception as e:
            results[account_id][region]['EC2'] = [{"error": str(e)}]

        try:
            efs = boto3.client('efs', region_name=region, **creds)
            filesystems = efs.describe_file_systems()['FileSystems']
            fs_data = []
            total_size = 0
            for fs in filesystems:
                size_gb = fs.get("SizeInBytes", {}).get("Value", 0) / 1024**3
                total_size += size_gb
                fs_data.append({
                    'ID': fs['FileSystemId'],
                    'PerformanceMode': fs.get('PerformanceMode'),
                    'Size_GB': size_gb,
                    'Tags': fs.get('Tags', [])
                })
            results[account_id][region]['EFS'] = fs_data
            summary.append([account_id, region, 'EFS', len(fs_data), round(total_size, 2)])
        except Exception as e:
            results[account_id][region]['EFS'] = [{"error": str(e)}]

        try:
            rds = boto3.client('rds', region_name=region, **creds)
            dbs = rds.describe_db_instances()['DBInstances']
            db_data = []
            total_size = 0
            for db in dbs:
                size = db['AllocatedStorage']
                total_size += size
                db_data.append({
                    'DBIdentifier': db['DBInstanceIdentifier'],
                    'Class': db['DBInstanceClass'],
                    'Allocated_Storage_GB': size,
                    'Engine': db['Engine']
                })
            results[account_id][region]['RDS'] = db_data
            summary.append([account_id, region, 'RDS', len(db_data), total_size])
        except Exception as e:
            results[account_id][region]['RDS'] = [{"error": str(e)}]

    try:
        s3 = boto3.client('s3', **creds)
        cw = boto3.client('cloudwatch', **creds)
        buckets = s3.list_buckets()['Buckets']
        bucket_data = []
        total_size = 0
        for bucket in buckets:
            try:
                name = bucket['Name']
                loc = s3.get_bucket_location(Bucket=name)['LocationConstraint'] or 'us-east-1'
                metrics = cw.get_metric_statistics(
                    Namespace='AWS/S3',
                    MetricName='BucketSizeBytes',
                    Dimensions=[
                        {'Name': 'BucketName', 'Value': name},
                        {'Name': 'StorageType', 'Value': 'StandardStorage'}
                    ],
                    StartTime=datetime.datetime.utcnow() - datetime.timedelta(days=3),
                    EndTime=datetime.datetime.utcnow(),
                    Period=86400,
                    Statistics=['Average']
                )
                size_bytes = metrics['Datapoints'][0]['Average'] if metrics['Datapoints'] else 0
                size_gb = size_bytes / 1024**3
                total_size += size_gb
                bucket_data.append({'Name': name, 'Region': loc, 'Size_GB': size_gb})
            except:
                bucket_data.append({'Name': name, 'Region': "unknown", 'Size_GB': 0})
        results[account_id]['global']['S3'] = bucket_data
        summary.append([account_id, 'global', 'S3', len(bucket_data), round(total_size, 2)])
    except Exception as e:
        results[account_id]['global']['S3'] = [{"error": str(e)}]

def write_excel(results, summary, filename="aws_inventory.xlsx"):
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Account ID", "Region", "Service", "Resource Count", "Total Capacity (GB)"])
    for row in summary:
        ws_summary.append(row)
    for ws in [ws_summary]:
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
    for account_id, regions in results.items():
        for region, services in regions.items():
            for service, items in services.items():
                sheet_name = f"{service}_{account_id[-4:]}_{region.replace('-', '')}"
                ws = wb.create_sheet(title=sheet_name[:31])
                if items and isinstance(items, list):
                    keys = set().union(*(d.keys() for d in items if isinstance(d, dict)))
                    ws.append(list(keys))
                    for item in items:
                        ws.append([item.get(k, "") for k in keys])
    wb.save(filename)

for acct in accounts:
    if acct['Status'] != 'ACTIVE':
        continue
    account_id = acct['Id']
    print(f"→ Collecting from Account: {account_id}")
    creds = assume_role(account_id)
    if creds:
        collect_inventory(account_id, creds)

print(json.dumps(results, indent=2))
write_excel(results, summary)
